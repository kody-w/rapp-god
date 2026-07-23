"""SpineDAG — universal Directed Acyclic Graph analyzer.

Point it at anything (a folder of code, an Excel workbook, a JSON/YAML
config, even a chunk of prose) and it will:

    * detect the input type and pick the right parser
    * build a node + edge graph
    * compute hubs, leaves, orphans, depth, cycles, and the critical (longest) path
    * score the graph's overall **Spine Health**
    * emit a Mermaid diagram, an ASCII tree, and a structured JSON envelope

The JSON envelope is delimited by `<spine-dag-json>...</spine-dag-json>`
so a UI (such as the bundled `spine_dag_demo.html`) can pull it out of the
brainstem chat response and render the visualisation.

No external dependencies for code/json/yaml/folder modes (stdlib only).
The Excel mode lazy-imports openpyxl when needed.
"""
from __future__ import annotations

import ast
import json
import os
import re
import sys
import time
from collections import defaultdict, deque

try:
    from basic_agent import BasicAgent  # local brainstem
except ImportError:  # cloud / openrappter
    from openrappter.agents.basic_agent import BasicAgent  # type: ignore


__manifest__ = {
    "schema": "rapp-agent/1.0",
    "name": "@rapp/spine_dag",
    "display_name": "SpineDAG",
    "description": (
        "Universal Directed-Acyclic-Graph analyser. Point it at any folder, "
        "code project, Excel workbook, JSON/YAML config, or block of prose; "
        "it builds a dependency graph, finds hubs, leaves, orphans, cycles, "
        "and the critical path, scores Spine Health 0-100, and returns a "
        "Markdown report with Mermaid diagram, ASCII tree, and a "
        "<spine-dag-json>...</spine-dag-json> envelope a UI can render."
    ),
    "author": "RAPP",
    "version": "1.0.0",
    "tags": ["dag", "graph", "dependency", "audit", "mermaid", "analysis",
             "excel", "code", "visualisation"],
    "category": "analysis",
    "quality_tier": "official",
    "requires_env": [],
    "dependencies": ["@rapp/basic_agent"],
    "example_call": {"args": {"target": "C:\\path\\to\\project", "mode": "auto"}},
}


# =====================================================================
# Tiny graph engine
# =====================================================================

class DAG:
    """Minimal directed graph with metric helpers. Cycles are tolerated
    (we are *not* an acyclic-only graph) — we detect & report them."""

    __slots__ = ("nodes", "edges", "labels", "node_kind")

    def __init__(self) -> None:
        self.nodes: set[str] = set()
        self.edges: dict[tuple[str, str], int] = defaultdict(int)
        self.labels: dict[str, str] = {}
        self.node_kind: dict[str, str] = {}

    # --- mutation ---------------------------------------------------
    def add_node(self, n: str, label: str | None = None, kind: str = "node") -> None:
        self.nodes.add(n)
        self.labels.setdefault(n, label or n)
        self.node_kind.setdefault(n, kind)

    def add_edge(self, src: str, dst: str, weight: int = 1) -> None:
        if not src or not dst or src == dst:
            return
        self.add_node(src); self.add_node(dst)
        self.edges[(src, dst)] += weight

    # --- views ------------------------------------------------------
    def out_neighbors(self, n: str) -> list[str]:
        return [d for (s, d) in self.edges if s == n]

    def in_neighbors(self, n: str) -> list[str]:
        return [s for (s, d) in self.edges if d == n]

    def in_deg(self) -> dict[str, int]:
        d = {n: 0 for n in self.nodes}
        for (_, dst), w in self.edges.items():
            d[dst] += w
        return d

    def out_deg(self) -> dict[str, int]:
        d = {n: 0 for n in self.nodes}
        for (src, _), w in self.edges.items():
            d[src] += w
        return d

    # --- metrics ----------------------------------------------------
    def find_cycles(self, max_cycles: int = 6) -> list[list[str]]:
        """Find simple cycles via Tarjan-ish DFS. Capped to keep things fast."""
        cycles: list[list[str]] = []
        adj: dict[str, list[str]] = defaultdict(list)
        for (s, d) in self.edges:
            adj[s].append(d)

        color: dict[str, int] = {n: 0 for n in self.nodes}  # 0=white,1=gray,2=black
        parent: dict[str, str | None] = {n: None for n in self.nodes}

        def dfs(u: str) -> None:
            if len(cycles) >= max_cycles:
                return
            color[u] = 1
            for v in adj.get(u, []):
                if len(cycles) >= max_cycles:
                    return
                if color[v] == 0:
                    parent[v] = u
                    dfs(v)
                elif color[v] == 1:
                    # back-edge -> cycle from v..u + v
                    path = [u]; cur = u
                    while cur is not None and cur != v:
                        cur = parent[cur]
                        if cur is not None:
                            path.append(cur)
                    path.reverse()
                    if path and path[0] == v:
                        cycles.append(path + [v])
            color[u] = 2

        sys.setrecursionlimit(max(10000, sys.getrecursionlimit()))
        for n in list(self.nodes):
            if color[n] == 0 and len(cycles) < max_cycles:
                try:
                    dfs(n)
                except RecursionError:
                    break
        return cycles

    def topo_layers(self) -> tuple[list[list[str]], set[str]]:
        """Kahn's algorithm. Returns (layers, nodes_in_cycles)."""
        indeg = {n: 0 for n in self.nodes}
        for (_, d) in self.edges:
            indeg[d] += 1
        layers: list[list[str]] = []
        ready = deque(sorted(n for n, d in indeg.items() if d == 0))
        seen = set()
        while ready:
            layer = []
            for _ in range(len(ready)):
                n = ready.popleft(); seen.add(n); layer.append(n)
                for v in self.out_neighbors(n):
                    indeg[v] -= 1
                    if indeg[v] == 0:
                        ready.append(v)
            layers.append(sorted(layer))
        in_cycle = self.nodes - seen
        return layers, in_cycle

    def critical_path(self) -> list[str]:
        """Longest path through the DAG (ignoring nodes in cycles)."""
        layers, _ = self.topo_layers()
        order = [n for layer in layers for n in layer]
        dist = {n: 0 for n in order}
        prev: dict[str, str | None] = {n: None for n in order}
        for u in order:
            for v in self.out_neighbors(u):
                if v not in dist:
                    continue
                if dist[u] + 1 > dist[v]:
                    dist[v] = dist[u] + 1
                    prev[v] = u
        if not dist:
            return []
        end = max(dist, key=dist.get)
        path = [end]
        while prev.get(path[-1]) is not None:
            path.append(prev[path[-1]])  # type: ignore[arg-type]
        return list(reversed(path))


# =====================================================================
# Parsers (one per mode)
# =====================================================================

_PY_FILE_RE = re.compile(r"\.py$", re.IGNORECASE)
_JS_FILE_RE = re.compile(r"\.(?:js|jsx|ts|tsx|mjs|cjs)$", re.IGNORECASE)
_JS_IMPORT_RE = re.compile(
    r"""(?:^|\n)\s*(?:import\s+(?:[^'";]+\s+from\s+)?|require\s*\(\s*)
        ['"]([^'"]+)['"]""", re.VERBOSE)


def _py_imports(path: str) -> list[str]:
    try:
        with open(path, "r", encoding="utf-8", errors="replace") as f:
            src = f.read()
        tree = ast.parse(src, filename=path)
    except Exception:
        return []
    out: list[str] = []
    for node in ast.walk(tree):
        if isinstance(node, ast.Import):
            for alias in node.names:
                out.append(alias.name.split(".", 1)[0])
        elif isinstance(node, ast.ImportFrom) and node.module:
            out.append(node.module.split(".", 1)[0])
    return out


def parse_python_dir(root: str, max_files: int = 600) -> tuple[DAG, dict]:
    g = DAG()
    files = []
    for dirpath, dirs, filenames in os.walk(root):
        # skip noise
        dirs[:] = [d for d in dirs if d not in (
            "__pycache__", ".git", ".venv", "venv", "node_modules",
            ".tox", "build", "dist", ".mypy_cache", ".pytest_cache")]
        for fn in filenames:
            if _PY_FILE_RE.search(fn):
                files.append(os.path.join(dirpath, fn))
        if len(files) > max_files:
            break
    files = files[:max_files]

    # Build module index by stem (best-effort intra-project resolution)
    stems = {os.path.splitext(os.path.basename(f))[0]: f for f in files}
    pkgs = set()
    for f in files:
        rel = os.path.relpath(f, root).replace(os.sep, ".")
        rel = rel[:-3]  # strip .py
        pkgs.add(rel.split(".", 1)[0])

    for f in files:
        node = os.path.relpath(f, root).replace(os.sep, "/")
        g.add_node(node, label=node, kind="module")
    for f in files:
        src_node = os.path.relpath(f, root).replace(os.sep, "/")
        for imp in _py_imports(f):
            if imp in stems and stems[imp] != f:
                tgt = os.path.relpath(stems[imp], root).replace(os.sep, "/")
                g.add_edge(src_node, tgt)
            elif imp in pkgs:
                # Internal package import; create a synthetic node
                pkg_node = f"pkg:{imp}"
                g.add_node(pkg_node, label=imp, kind="package")
                g.add_edge(src_node, pkg_node)

    return g, {"language": "python", "files_scanned": len(files), "root": root}


def parse_js_dir(root: str, max_files: int = 600) -> tuple[DAG, dict]:
    g = DAG(); files: list[str] = []
    for dirpath, dirs, filenames in os.walk(root):
        dirs[:] = [d for d in dirs if d not in (
            "node_modules", ".git", "build", "dist", ".next", "out", ".turbo")]
        for fn in filenames:
            if _JS_FILE_RE.search(fn):
                files.append(os.path.join(dirpath, fn))
        if len(files) > max_files:
            break
    files = files[:max_files]

    file_set = {os.path.normpath(f): f for f in files}
    for f in files:
        n = os.path.relpath(f, root).replace(os.sep, "/")
        g.add_node(n, label=n, kind="module")

    for f in files:
        src_node = os.path.relpath(f, root).replace(os.sep, "/")
        try:
            with open(f, "r", encoding="utf-8", errors="replace") as fh:
                src = fh.read()
        except Exception:
            continue
        for m in _JS_IMPORT_RE.finditer(src):
            spec = m.group(1)
            if spec.startswith("."):
                # resolve relative
                base = os.path.dirname(f)
                cand_paths = [
                    os.path.normpath(os.path.join(base, spec)),
                    os.path.normpath(os.path.join(base, spec + ".js")),
                    os.path.normpath(os.path.join(base, spec + ".ts")),
                    os.path.normpath(os.path.join(base, spec + ".jsx")),
                    os.path.normpath(os.path.join(base, spec + ".tsx")),
                    os.path.normpath(os.path.join(base, spec, "index.js")),
                    os.path.normpath(os.path.join(base, spec, "index.ts")),
                ]
                for c in cand_paths:
                    if c in file_set:
                        tgt = os.path.relpath(c, root).replace(os.sep, "/")
                        g.add_edge(src_node, tgt); break
            else:
                pkg_node = f"pkg:{spec.split('/', 1)[0]}"
                g.add_node(pkg_node, label=spec.split('/', 1)[0], kind="package")
                g.add_edge(src_node, pkg_node)

    return g, {"language": "javascript", "files_scanned": len(files), "root": root}


def parse_excel(path: str) -> tuple[DAG, dict]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise RuntimeError("openpyxl not installed; install it to analyse spreadsheets")
    g = DAG()
    wb = load_workbook(path, read_only=True, data_only=False, keep_vba=False)
    sheets = wb.sheetnames
    sheet_set = {s.lower() for s in sheets}
    for s in sheets:
        g.add_node(s, label=s, kind="sheet")
    pat = re.compile(r"(?:'([^']+)'|([A-Za-z0-9_.\-]+))!")
    formula_count = 0
    sampled = 0
    for s in sheets:
        ws = wb[s]
        for row in ws.iter_rows(values_only=False):
            for cell in row:
                v = cell.value
                if isinstance(v, str) and v.startswith("="):
                    formula_count += 1
                    if sampled >= 10000:
                        continue
                    sampled += 1
                    for m in pat.finditer(v):
                        tgt = (m.group(1) or m.group(2)).strip()
                        if tgt.lower() in sheet_set and tgt.lower() != s.lower():
                            g.add_edge(s, tgt)
    wb.close()
    return g, {"language": "excel", "sheets": len(sheets),
               "formulas": formula_count, "sampled": sampled, "root": path}


def parse_json_yaml(path: str, max_depth: int = 8) -> tuple[DAG, dict]:
    """Object reference graph: each $ref or {key} placeholder becomes an edge."""
    text = open(path, "r", encoding="utf-8", errors="replace").read()
    if path.lower().endswith((".yaml", ".yml")):
        try:
            import yaml  # type: ignore
            data = yaml.safe_load(text)
            kind = "yaml"
        except Exception:
            data = None; kind = "yaml"
    else:
        try:
            data = json.loads(text); kind = "json"
        except Exception:
            data = None; kind = "json"
    g = DAG()
    if data is None:
        return g, {"language": kind, "root": path, "parse_error": True}

    def walk(node, path_str: str, depth: int = 0):
        if depth > max_depth: return
        if isinstance(node, dict):
            for k, v in node.items():
                child = f"{path_str}.{k}" if path_str else k
                g.add_node(child, label=k, kind="key")
                if path_str:
                    g.add_edge(path_str, child)
                # $ref-style edges
                if isinstance(v, str) and (v.startswith("#/") or "$ref" == k.lower()):
                    target = v.lstrip("#/")
                    g.add_node(target, label=target, kind="ref"); g.add_edge(child, target)
                walk(v, child, depth + 1)
        elif isinstance(node, list):
            for i, v in enumerate(node):
                walk(v, f"{path_str}[{i}]", depth + 1)
    walk(data, "")
    return g, {"language": kind, "root": path}


def parse_generic_folder(root: str, max_files: int = 400) -> tuple[DAG, dict]:
    """Filename mention graph: which file names appear in which other file."""
    g = DAG()
    files = []
    for dirpath, dirs, filenames in os.walk(root):
        dirs[:] = [d for d in dirs if not d.startswith(".") and d not in (
            "node_modules", "__pycache__", "venv", ".venv", "build", "dist")]
        for fn in filenames:
            if fn.startswith("."): continue
            files.append(os.path.join(dirpath, fn))
        if len(files) > max_files:
            break
    files = files[:max_files]
    names = {os.path.basename(f): f for f in files}
    for f in files:
        rel = os.path.relpath(f, root).replace(os.sep, "/")
        g.add_node(rel, label=rel, kind="file")
    for f in files:
        src_rel = os.path.relpath(f, root).replace(os.sep, "/")
        try:
            with open(f, "r", encoding="utf-8", errors="replace") as fh:
                txt = fh.read(200_000)
        except Exception:
            continue
        for name, fp in names.items():
            if name == os.path.basename(f) or name not in txt:
                continue
            tgt_rel = os.path.relpath(fp, root).replace(os.sep, "/")
            g.add_edge(src_rel, tgt_rel)
    return g, {"language": "generic", "files_scanned": len(files), "root": root}


def parse_text(text: str) -> tuple[DAG, dict]:
    """Heuristic relation extractor: finds `A -> B`, `A depends on B`, etc."""
    g = DAG()
    arrow_re = re.compile(r"\b([A-Z][A-Za-z0-9_ -]{1,40}?)\s*(?:->|→|=>|leads to|drives|depends on|requires|feeds)\s*([A-Z][A-Za-z0-9_ -]{1,40})\b")
    for m in arrow_re.finditer(text or ""):
        a = m.group(1).strip(); b = m.group(2).strip()
        g.add_edge(a, b)
    return g, {"language": "text", "edges_extracted": len(g.edges)}


# =====================================================================
# Mode dispatch
# =====================================================================

def detect_mode(path: str) -> str:
    if not path:
        return "text"
    p = path.lower()
    if os.path.isfile(path):
        if p.endswith((".xlsx", ".xlsm", ".xlsb")): return "excel"
        if p.endswith((".json",)):                  return "json"
        if p.endswith((".yaml", ".yml")):           return "yaml"
        return "generic-file"
    if os.path.isdir(path):
        # heuristic: any .py?
        for root_, _dirs, files in os.walk(path):
            if any(f.endswith(".py") for f in files):
                return "python"
            if any(f.endswith((".js",".ts",".jsx",".tsx")) for f in files):
                return "javascript"
            break
        return "generic-folder"
    return "text"


def run_mode(mode: str, path_or_text: str) -> tuple[DAG, dict]:
    if mode == "python":         return parse_python_dir(path_or_text)
    if mode == "javascript":     return parse_js_dir(path_or_text)
    if mode == "excel":          return parse_excel(path_or_text)
    if mode in ("json","yaml"):  return parse_json_yaml(path_or_text)
    if mode == "generic-folder": return parse_generic_folder(path_or_text)
    if mode == "generic-file":
        # treat as text-of-file
        try:
            return parse_text(open(path_or_text, "r", encoding="utf-8", errors="replace").read())
        except Exception:
            return DAG(), {"language": "text", "error": "cannot read"}
    return parse_text(path_or_text)


# =====================================================================
# Render helpers
# =====================================================================

def safe_id(s: str) -> str:
    return "n_" + re.sub(r"[^A-Za-z0-9]", "_", s)[:48]


def render_mermaid(g: DAG, max_edges: int = 60) -> str:
    if not g.edges:
        return 'graph LR\n  empty["(no edges)"]'
    by_w = sorted(g.edges.items(), key=lambda kv: -kv[1])[:max_edges]
    seen = {}
    lines = ["graph LR"]
    in_d = g.in_deg()
    cycles = g.find_cycles(max_cycles=4)
    cycle_set = {n for c in cycles for n in c}
    for (s, d), w in by_w:
        sid = safe_id(s); did = safe_id(d)
        seen[sid] = (s, in_d.get(s, 0))
        seen[did] = (d, in_d.get(d, 0))
        arrow = "==>" if (s in cycle_set and d in cycle_set) else "-->"
        lines.append(f'  {sid}["{g.labels.get(s,s)}"] {arrow}|{w}| {did}["{g.labels.get(d,d)}"]')
    # class hubs
    hub_ids = [safe_id(n) for n, _ in sorted(in_d.items(), key=lambda kv: -kv[1])[:3]]
    if hub_ids:
        lines.append("  classDef hub fill:#7c5cfc,stroke:#a78bfa,color:#fff")
        lines.append("  class " + ",".join(hub_ids) + " hub")
    if cycle_set:
        cyc_ids = ",".join({safe_id(n) for n in cycle_set})
        lines.append("  classDef cyc fill:#fb7185,stroke:#fda4af,color:#fff")
        lines.append("  class " + cyc_ids + " cyc")
    return "\n".join(lines)


def render_ascii_tree(g: DAG, max_lines: int = 40) -> str:
    layers, in_cycle = g.topo_layers()
    out: list[str] = []
    for li, layer in enumerate(layers):
        prefix = "│  " * li
        for n in layer[:6]:
            tag = "★" if n in in_cycle else "•"
            out.append(f"{prefix}{tag} {g.labels.get(n,n)}")
            if len(out) >= max_lines: return "\n".join(out) + "\n  …(truncated)"
        if len(layer) > 6:
            out.append(f"{prefix}  …(+{len(layer)-6} more)")
    return "\n".join(out) or "(empty)"


def health_score(g: DAG, cycles: list[list[str]], orphans: list[str]) -> tuple[int, str]:
    """0-100. Higher = healthier spine."""
    if not g.nodes:
        return 0, "empty"
    n = len(g.nodes); e = sum(g.edges.values())
    score = 100
    score -= min(40, len(cycles) * 12)         # cycles bad
    score -= min(20, int(100 * len(orphans) / max(1, n)))  # too many orphans
    # fan-out blowup
    in_d = g.in_deg(); top = max(in_d.values()) if in_d else 0
    if top > n * 0.4:
        score -= 15  # godfile penalty
    # too few edges
    if e < n // 4:
        score -= 10
    score = max(0, min(100, score))
    band = "Excellent" if score >= 85 else "Healthy" if score >= 70 else "Fragile" if score >= 45 else "Tangled"
    return score, band


def build_report(mode: str, target: str, g: DAG, meta: dict, elapsed: float) -> str:
    in_d = g.in_deg(); out_d = g.out_deg()
    total_edges = sum(g.edges.values())
    hubs = sorted(in_d.items(), key=lambda kv: -kv[1])[:5]
    leaves = sorted(out_d.items(), key=lambda kv: -kv[1])[:5]
    orphans = [n for n in g.nodes
               if in_d.get(n, 0) == 0 and out_d.get(n, 0) == 0]
    cycles = g.find_cycles(max_cycles=6)
    crit = g.critical_path()
    layers, in_cycle = g.topo_layers()
    score, band = health_score(g, cycles, orphans)

    mermaid = render_mermaid(g)
    tree = render_ascii_tree(g)

    envelope = {
        "schema": "spine-dag/1",
        "mode": mode,
        "target": target,
        "meta": meta,
        "metrics": {
            "nodes": len(g.nodes),
            "edges": total_edges,
            "edge_pairs": len(g.edges),
            "max_depth": len(layers),
            "orphans": len(orphans),
            "cycles": len(cycles),
            "in_cycle_nodes": len(in_cycle),
            "critical_path_len": len(crit),
            "health_score": score,
            "health_band": band,
        },
        "hubs":   [{"node": n, "label": g.labels.get(n, n), "in_refs":  c} for n, c in hubs],
        "leaves": [{"node": n, "label": g.labels.get(n, n), "out_refs": c} for n, c in leaves],
        "orphans": orphans[:30],
        "cycles": [[g.labels.get(n, n) for n in c] for c in cycles],
        "critical_path": [g.labels.get(n, n) for n in crit],
        "layers": [[g.labels.get(n, n) for n in layer] for layer in layers[:10]],
        "mermaid": mermaid,
        "elapsed_s": round(elapsed, 3),
    }

    def fmt_list(items): return ", ".join(items) if items else "—"
    score_bar = ("█" * (score // 5)) + ("░" * (20 - score // 5))

    md = f"""# 🧬 Spine DAG Analysis

**Target:** `{target}`  
**Mode detected:** `{mode}` · **Parsed in {elapsed:.2f}s**

## Spine Health: {score}/100 — *{band}*

```
{score_bar}  {score}/100
```

| Metric | Value |
|---|---|
| Nodes | **{len(g.nodes)}** |
| Edges | **{total_edges}** ({len(g.edges)} unique pairs) |
| Max depth | **{len(layers)}** layers |
| Orphans | {len(orphans)} |
| Cycles | {len(cycles)}{' ⚠️' if cycles else ''} |
| Critical path length | {len(crit)} nodes |

## 🌟 Top hubs (most depended-on)
{chr(10).join(f"- **{g.labels.get(n,n)}** — {c} incoming" for n,c in hubs) or '_(none)_'}

## 🍃 Top leaves (most outgoing)
{chr(10).join(f"- **{g.labels.get(n,n)}** — {c} outgoing" for n,c in leaves) or '_(none)_'}

## 🛤️ Critical path
{' → '.join(g.labels.get(n,n) for n in crit) if crit else '_(none — graph empty or fully cyclic)_'}

## 🔁 Cycles {'detected' if cycles else 'clean (none)'}
{chr(10).join('- ' + ' → '.join(g.labels.get(n,n) for n in c) for c in cycles) or '_(no cycles — pure DAG)_'}

## 🕳️ Orphans ({len(orphans)})
{fmt_list([g.labels.get(n,n) for n in orphans[:12]])}{ '  …' if len(orphans) > 12 else ''}

## 🌳 Topological layers
```
{tree}
```

## 🧭 Mermaid graph
```mermaid
{mermaid}
```

---
<spine-dag-json>
{json.dumps(envelope, separators=(',', ':'))}
</spine-dag-json>
"""
    return md


# =====================================================================
# Agent
# =====================================================================

class SpineDAGAgent(BasicAgent):
    def __init__(self) -> None:
        self.name = "SpineDAG"
        self.metadata = {
            "name": self.name,
            "description": (
                "Universal Directed-Acyclic-Graph analyser. Point it at any "
                "folder, code project, Excel workbook, JSON/YAML config, or "
                "block of prose and it builds a dependency graph, finds hubs, "
                "leaves, orphans, cycles and the critical path, scores the "
                "graph's Spine Health (0-100), and returns a Markdown report "
                "with an embedded Mermaid diagram, ASCII tree, and a "
                "<spine-dag-json>...</spine-dag-json> envelope a UI can render. "
                "Use whenever the user asks to map dependencies, audit "
                "structure, find circular references, or visualise how the "
                "pieces of something connect."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "target": {
                        "type": "string",
                        "description": "An absolute file path, folder path, or "
                                       "block of text to analyse. If omitted, "
                                       "the tool will fall back to the "
                                       "`text` parameter.",
                    },
                    "mode": {
                        "type": "string",
                        "enum": ["auto", "python", "javascript", "excel",
                                 "json", "yaml", "generic-folder",
                                 "generic-file", "text"],
                        "description": "Force a specific parser. Default `auto` "
                                       "detects from the target.",
                    },
                    "text": {
                        "type": "string",
                        "description": "Raw text to analyse when `target` is not "
                                       "a path (e.g. `A -> B; A depends on C`).",
                    },
                },
                "required": [],
            },
        }
        super().__init__(name=self.name, metadata=self.metadata)

    def perform(self, **kwargs) -> str:
        target = (kwargs.get("target") or "").strip()
        mode = (kwargs.get("mode") or "auto").strip().lower()
        text = kwargs.get("text") or ""

        if not target and text:
            mode = "text"
            payload = text
        elif target and (os.path.isfile(target) or os.path.isdir(target)):
            payload = target
            if mode == "auto":
                mode = detect_mode(target)
        elif target:
            mode = "text"; payload = target  # treat target as inline text
        else:
            return ("❌ SpineDAG needs either `target` (a path) or `text` "
                    "(a block of relations like `A -> B`).")

        t0 = time.time()
        try:
            g, meta = run_mode(mode, payload)
        except Exception as e:
            return f"❌ SpineDAG failed in mode `{mode}`: {e}"
        elapsed = time.time() - t0
        target_label = target if target else f"<{len(text)} chars of text>"
        return build_report(mode, target_label, g, meta, elapsed)
